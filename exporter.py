import json
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


def export_from_json_to_excel():
    # Пути к файлам
    excel_file_path = 'REGULAR_SCHEDULE.xlsx'
    filled_template_path = 'output_data/filled_template.json'
    priority_boost_data_path = 'output_data/priority_boost_data.json'

    # Загрузка данных из priority_boost_data.json
    with open(priority_boost_data_path, 'r') as f:
        boost_data = json.load(f)

    # Извлечение уникальных моделей
    unique_models = list(set(entry['Model'] for entry in boost_data))

    # Генерация цветов для каждой категории
    category_colors = {
        'priority': "8ED7DD",
        'boost': "FFFFFF",
        'vanila': "FFFF00",
        'bdsm': "FF0000",
        'cosplay': "FF00FF",
        'ebony': "7D5F02",
    }

    # Загружаем данные из JSON
    with open(filled_template_path, 'r') as f:
        filled_template_data = json.load(f)

    # Открываем файл Excel
    wb = openpyxl.load_workbook(excel_file_path)
    template_sheet = wb['Template']

    # Создаем копию листа Template с новым именем
    import time

    new_sheet_name = f"Schedule_{time.strftime('%Y%m%d_%H%M%S')}"
    wb.copy_worksheet(template_sheet).title = new_sheet_name
    new_sheet = wb[new_sheet_name]

    # Функция для нахождения начала и конца секций дней недели
    def find_day_section(sheet, day):
        for row in sheet.iter_rows(min_row=1, max_col=1):
            if row[0].value == day:
                start_row = row[0].row
                end_row = start_row + 6  # День занимает 7 строк
                return start_row, end_row
        return None, None

    # Проходим по всем моделям и заполняем данные
    for model_data in filled_template_data:
        model_name = model_data['Model']

        for day_data in model_data['Days']:
            day_name = day_data['Day'].upper()
            tasks = day_data['Filled Tasks']

            start_row, end_row = find_day_section(new_sheet, day_name)

            if start_row is None:
                continue

            # Найдем колонку с текущей моделью
            for col_idx, col in enumerate(new_sheet.iter_cols(min_row=1, max_row=1), start=1):
                if col[0].value == model_name:
                    column_letter = get_column_letter(col_idx+1)

                    # Заполняем задачи
                    for i, task in enumerate(tasks, start=start_row):
                        cell = new_sheet[f"{column_letter}{i}"]

                        # Определяем категорию и очищаем имя
                        if task.endswith('_priority'):
                            base_name = task.rstrip('_priority')
                            cell.fill = PatternFill(start_color=category_colors['priority'], end_color=category_colors['priority'], fill_type="solid")
                        elif task.endswith('_boost'):
                            base_name = task.rstrip('_boost')
                            cell.fill = PatternFill(start_color=category_colors['boost'], end_color=category_colors['boost'], fill_type="solid")
                        elif task.endswith('_vanila'):
                            base_name = task.rstrip('_vanila')
                            cell.fill = PatternFill(start_color=category_colors['vanila'], end_color=category_colors['vanila'], fill_type="solid")
                        elif task.endswith('_bdsm'):
                            base_name = task.rstrip('_bdsm')
                            cell.fill = PatternFill(start_color=category_colors['bdsm'], end_color=category_colors['bdsm'], fill_type="solid")
                        elif task.endswith('_cosplay'):
                            base_name = task.rstrip('_cosplay')
                            cell.fill = PatternFill(start_color=category_colors['cosplay'], end_color=category_colors['cosplay'], fill_type="solid")
                        elif task.endswith('_ebony'):
                            base_name = task.rstrip('_ebony')
                            cell.fill = PatternFill(start_color=category_colors['ebony'], end_color=category_colors['ebony'], fill_type="solid")
                        else:
                            base_name = task

                        # Применяем жирный шрифт
                        cell.font = Font(bold=True)
                        cell.value = base_name

                    break

    # Сохраняем файл Excel
    wb.save(excel_file_path)
    print(f"Данные успешно заполнены и сохранены в файл: {excel_file_path}")