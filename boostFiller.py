"""import json
import random
import os
import openpyxl


def fill_boost():
    # Пути к основным файлам
    models_data_path = 'output_data/models_data.json'
    blacklist_file_path = 'output_data/blacklist_data.json'
    schedule_dir = 'output_data/days_split'
    boost_dir = 'output_data/boost_days_split'
    excel_file_path = 'REGULAR_SCHEDULE.xlsx'

    # Загрузка общих данных
    with open(models_data_path, 'r') as f:
        models_data = json.load(f)

    with open(blacklist_file_path, 'r') as f:
        blacklist_data = json.load(f)

    # Преобразование блокировок для удобства поиска
    blocked_dict = {item['Model']: item['Blocked'] for item in blacklist_data}

    # Создание словаря с категориями моделей
    model_to_category = {model['Model']: model['Category'] for model in models_data}

    # Определение приоритетов для каждой категории
    priority_order = {
        'VANILA': ['VANILA', 'COSPLAY', 'BDSM', 'EBONY'],
        'COSPLAY': ['COSPLAY', 'VANILA', 'BDSM', 'EBONY'],
        'BDSM': ['BDSM', 'COSPLAY', 'VANILA', 'EBONY'],
        'EBONY': ['EBONY', 'VANILA', 'COSPLAY', 'BDSM']
    }

    # Функция для проверки блокировок
    def is_blocked(model, target_model):
        blocked_models_for_target = blocked_dict.get(target_model, [])
        return model in blocked_models_for_target

    # Функция для распределения бустов по пулу моделей
    def distribute_boost_in_pool(model_name, boost_count, pool_models, tasks_dict):
        for entry in pool_models:
            tasks = tasks_dict[entry['Model']]
            available_slots = [i for i, task in enumerate(tasks) if task == 'BOOST']

            for slot in available_slots:
                if (
                        #TODO
                        #Проверка на разрешоную категорию
                        #- Проверка категории таргет модели
                        boost_count > 0
                        and model_name not in tasks
                        and not is_blocked(model_name, entry['Model'])
                        and model_name != entry['Model']
                ):
                    tasks[slot] = f"{model_name}_boost"
                    boost_count -= 1
                    break  # Нам нужно только первое подходящее место

            if boost_count == 0:
                break

        return boost_count

    # Функция для распределения бустов с использованием пула
    def distribute_boosts(schedule_file_path, boost_file_path):
        with open(schedule_file_path, 'r') as f:
            day_schedule = json.load(f)

        with open(boost_file_path, 'r') as f:
            day_boost = json.load(f)

        schedule_dict = {entry['Model']: entry['Tasks'] for entry in day_schedule}
        boost_pool = day_boost.copy()

        remaining_boosts = []

        while boost_pool:
            random_entry = random.choice(boost_pool)
            model_name = random_entry['Model']
            boost_count = random_entry['Boost Count']
            model_category = model_to_category[model_name]
            categories_to_check = priority_order[model_category]

            for category in categories_to_check:
                if boost_count <= 0:
                    break

                pool_models = [entry for entry in day_schedule if model_to_category[entry['Model']] == category]

                boost_count = distribute_boost_in_pool(model_name, boost_count, pool_models, schedule_dict)

            random_entry['Boost Count'] = boost_count

            remaining_boosts.append({
                'Model': model_name,
                'Day': os.path.basename(schedule_file_path).split('_')[0].capitalize(),
                'Remaining Boost Count': boost_count
            })

            boost_pool.remove(random_entry)

        with open(schedule_file_path, 'w') as f:
            json.dump(day_schedule, f, indent=4)

        with open(boost_file_path, 'w') as f:
            json.dump(day_boost, f, indent=4)

        return remaining_boosts

    # Основная обработка по дням недели
    remaining_boosts_all_days = []

    for day in ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']:
        schedule_file = os.path.join(schedule_dir, f'{day}_schedule.json')
        boost_file = os.path.join(boost_dir, f'{day}_boost.json')
        remaining_boosts = distribute_boosts(schedule_file, boost_file)
        remaining_boosts_all_days.extend(remaining_boosts)

    # Открываем или создаем Excel файл
    wb = openpyxl.load_workbook(excel_file_path)
    boost_sheet = wb['Boost']

    # Записываем данные по остаткам бустов обратно в лист Boost
    days_list = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']

    for boost_entry in remaining_boosts_all_days:
        day = boost_entry['Day'].upper()
        day_col = days_list.index(day) + 1  # Находим индекс дня недели
        model_name = boost_entry['Model']
        remaining_boost_count = boost_entry['Remaining Boost Count']

        # Найти строку с нужной моделью
        for row in boost_sheet.iter_rows(min_row=2, max_row=boost_sheet.max_row):
            if row[0].value == model_name:
                # Записать остаток буста в нужную колонку
                row[day_col].value = remaining_boost_count
                break

    # Сохраняем файл Excel
    wb.save(excel_file_path)
    print(f"Данные по остаткам бустов сохранены в файл: {excel_file_path}")
"""

"""import json
import random
import os
import openpyxl


def fill_boost():
    # Пути к основным файлам
    models_data_path = 'output_data/models_data.json'
    blacklist_file_path = 'output_data/blacklist_data.json'
    schedule_dir = 'output_data/days_split'
    boost_dir = 'output_data/boost_days_split'
    excel_file_path = 'REGULAR_SCHEDULE.xlsx'

    # Загрузка общих данных
    with open(models_data_path, 'r') as f:
        models_data = json.load(f)

    with open(blacklist_file_path, 'r') as f:
        blacklist_data = json.load(f)

    # Преобразование блокировок для удобства поиска
    blocked_dict = {item['Model']: item['Blocked'] for item in blacklist_data}

    # Создание словаря с категориями моделей
    model_to_category = {model['Model']: model['Category'] for model in models_data}

    # Определение приоритетов для каждой категории
    priority_order = {
        'VANILA': ['VANILA', 'COSPLAY', 'BDSM', 'EBONY'],
        'COSPLAY': ['COSPLAY', 'VANILA', 'BDSM', 'EBONY'],
        'BDSM': ['BDSM', 'COSPLAY', 'VANILA', 'EBONY'],
        'EBONY': ['EBONY', 'VANILA', 'COSPLAY', 'BDSM']
    }

    # Функция для проверки наличия свободных ячеек BOOST и их количества
    def check_available_boost_slots(pool_models, tasks_dict):
        total_slots = 0
        for entry in pool_models:
            tasks = tasks_dict[entry['Model']]
            available_slots = sum(1 for task in tasks if task == 'BOOST')
            total_slots += available_slots
        return total_slots

    # Функция для проверки блокировок
    def is_blocked(model, target_model):
        blocked_models_for_target = blocked_dict.get(target_model, [])
        return model in blocked_models_for_target

    # Функция для распределения бустов по пулу моделей
    def distribute_boost_in_pool(model_name, boost_count, pool_models, tasks_dict):
        for entry in pool_models:
            tasks = tasks_dict[entry['Model']]
            available_slots = [i for i, task in enumerate(tasks) if task == 'BOOST']

            for slot in available_slots:
                if (
                        boost_count > 0
                        and model_name not in tasks
                        and f"{model_name}_boost" not in tasks
                        and not is_blocked(model_name, entry['Model'])
                        and model_name != entry['Model']
                ):
                    tasks[slot] = f"{model_name}_boost"
                    boost_count -= 1
                    break  # Нам нужно только первое подходящее место

            if boost_count == 0:
                break

        return boost_count

    # Функция для разделения boost_pool на подпулы
    def divide_boost_pool(boost_pool):
        # Определяем максимальное количество бустов у любой модели
        max_boosts = max(entry['Boost Count'] for entry in boost_pool)

        # Создаем список пустых подпулов
        subpools = [[] for _ in range(max_boosts)]

        # Распределяем модели по подпулам
        for entry in boost_pool:
            model_name = entry['Model']
            boost_count = entry['Boost Count']

            # Добавляем модель в каждый подпул
            for i in range(boost_count):
                subpools[i].append({'Model': model_name, 'Boost Count': 1})

        return subpools

    # Функция для распределения бустов с использованием подпулов
    def distribute_boosts(schedule_file_path, boost_file_path):
        with open(schedule_file_path, 'r') as f:
            day_schedule = json.load(f)

        with open(boost_file_path, 'r') as f:
            day_boost = json.load(f)

        schedule_dict = {entry['Model']: entry['Tasks'] for entry in day_schedule}

        # Разделяем boost_pool на подпулы
        boost_pool = day_boost.copy()
        subpools = divide_boost_pool(boost_pool)

        remaining_boosts = []

        # Проходим по каждому подпулу
        for subpool in subpools:
            while subpool:
                random_entry = random.choice(subpool)
                model_name = random_entry['Model']
                boost_count = random_entry['Boost Count']
                model_category = model_to_category[model_name]
                categories_to_check = priority_order[model_category]

                for category in categories_to_check:
                    if boost_count <= 0:
                        break

                    pool_models = [entry for entry in day_schedule if model_to_category[entry['Model']] == category]

                    boost_count = distribute_boost_in_pool(model_name, boost_count, pool_models, schedule_dict)

                if boost_count < random_entry['Boost Count']:  # Если буст был успешно использован
                    random_entry['Boost Count'] -= 1  # Уменьшаем счетчик бустов для текущей модели

                if random_entry['Boost Count'] == 0:
                    subpool.remove(random_entry)
                    # Обновляем оригинальный boost_pool
                    for entry in boost_pool:
                        if entry['Model'] == random_entry['Model']:
                            entry['Boost Count'] -= 1

        remaining_boosts.extend(
            {
                'Model': entry['Model'],
                'Day': os.path.basename(schedule_file_path).split('_')[0].capitalize(),
                'Remaining Boost Count': entry['Boost Count']
            }
            for entry in boost_pool if entry['Boost Count'] >= 0
        )

        with open(schedule_file_path, 'w') as f:
            json.dump(day_schedule, f, indent=4)

        with open(boost_file_path, 'w') as f:
            json.dump(day_boost, f, indent=4)

        return remaining_boosts

    # Основная обработка по дням недели
    remaining_boosts_all_days = []

    for day in ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']:
        schedule_file = os.path.join(schedule_dir, f'{day}_schedule.json')
        boost_file = os.path.join(boost_dir, f'{day}_boost.json')
        remaining_boosts = distribute_boosts(schedule_file, boost_file)
        remaining_boosts_all_days.extend(remaining_boosts)

    # Открываем или создаем Excel файл
    wb = openpyxl.load_workbook(excel_file_path)
    boost_sheet = wb['Boost']

    # Записываем данные по остаткам бустов обратно в лист Boost
    days_list = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']

    for boost_entry in remaining_boosts_all_days:
        day = boost_entry['Day'].upper()
        day_col = days_list.index(day) + 1  # Находим индекс дня недели
        model_name = boost_entry['Model']
        remaining_boost_count = boost_entry['Remaining Boost Count']

        # Найти строку с нужной моделью
        for row in boost_sheet.iter_rows(min_row=2, max_row=boost_sheet.max_row):
            if row[0].value == model_name:
                # Записать остаток буста в нужную колонку
                row[day_col].value = remaining_boost_count
                break

    # Сохраняем файл Excel
    wb.save(excel_file_path)
    print(f"Данные по остаткам бустов сохранены в файл: {excel_file_path}")
"""


import json
import random
import os
import openpyxl

def fill_boost():
    # Пути к основным файлам
    models_data_path = 'output_data/models_data.json'
    blacklist_file_path = 'output_data/blacklist_data.json'
    schedule_dir = 'output_data/days_split'
    boost_dir = 'output_data/boost_days_split'
    excel_file_path = 'REGULAR_SCHEDULE.xlsx'

    # Загрузка общих данных
    with open(models_data_path, 'r') as f:
        models_data = json.load(f)

    with open(blacklist_file_path, 'r') as f:
        blacklist_data = json.load(f)

    # Преобразование блокировок для удобства поиска
    blocked_dict = {item['Model']: item['Blocked'] for item in blacklist_data}

    # Создание словаря с категориями моделей
    model_to_category = {model['Model']: model['Category'] for model in models_data}

    # Определение приоритетов для каждой категории
    priority_order = {
        'VANILA': ['VANILA', 'COSPLAY', 'BDSM', 'EBONY'],
        'COSPLAY': ['COSPLAY', 'VANILA', 'BDSM', 'EBONY'],
        'BDSM': ['BDSM', 'COSPLAY', 'VANILA', 'EBONY'],
        'EBONY': ['EBONY', 'VANILA', 'COSPLAY', 'BDSM']
    }

    # Функция для проверки блокировок
    def is_blocked(model, target_model):
        blocked_models_for_target = blocked_dict.get(target_model, [])
        return model in blocked_models_for_target

    def check_available_boost_slots(day_schedule, tasks_dict):
        total_slots = 0
        for entry in day_schedule:
            tasks = tasks_dict[entry['Model']]
            for task in tasks:
                if task == "BOOST":
                    total_slots += 1
        print("Total available BOOST slots:", total_slots)
        return total_slots

    # Функция для распределения бустов по пулу моделей
    def distribute_boost_in_pool(model_name, boost_count, pool_models, tasks_dict, available_slots):
        for entry in pool_models:
            tasks = tasks_dict[entry['Model']]
            for slot_index, task in enumerate(tasks):
                if (
                        available_slots > 0 and
                        task == 'BOOST' and
                        model_name not in tasks and
                        f"{model_name}_boost" not in tasks and
                        not is_blocked(model_name, entry['Model']) and
                        model_name != entry['Model']
                ):
                    tasks[slot_index] = f"{model_name}_boost"
                    boost_count -= 1
                    available_slots -= 1
                    break  # Нам нужно только первое подходящее место

            if boost_count == 0 or available_slots == 0:
                break

        return boost_count, available_slots

    # Функция для разделения boost_pool на подпулы
    def divide_boost_pool(boost_pool):
        # Определяем максимальное количество бустов у любой модели
        max_boosts = max(entry['Boost Count'] for entry in boost_pool)

        # Создаем список пустых подпулов
        subpools = [[] for _ in range(max_boosts)]

        # Распределяем модели по подпулам
        for entry in boost_pool:
            model_name = entry['Model']
            boost_count = entry['Boost Count']

            # Добавляем модель в каждый подпул
            for i in range(boost_count):
                subpools[i].append({'Model': model_name, 'Boost Count': 1})

        return subpools

    # Функция для распределения бустов с использованием подпулов
    def distribute_boosts(schedule_file_path, boost_file_path):
        with open(schedule_file_path, 'r') as f:
            day_schedule = json.load(f)

        with open(boost_file_path, 'r') as f:
            day_boost = json.load(f)

        schedule_dict = {entry['Model']: entry['Tasks'] for entry in day_schedule}

        # Разделяем boost_pool на подпулы
        boost_pool = day_boost.copy()
        subpools = divide_boost_pool(boost_pool)

        remaining_boosts = []

        # Проверяем наличие свободных ячеек перед началом процесса
        available_slots = check_available_boost_slots(day_schedule, schedule_dict)
        if available_slots == 0:
            print("Нет доступных слотов BOOST для заполнения.")
            return remaining_boosts

        # Проходим по каждому подпулу
        for subpool in subpools:
            if available_slots == 0:
                break
            while subpool and available_slots != 0:
                random_entry = random.choice(subpool)
                model_name = random_entry['Model']
                boost_count = random_entry['Boost Count']
                model_category = model_to_category[model_name]
                categories_to_check = priority_order[model_category]

                for category in categories_to_check:
                    if boost_count == 0 or available_slots == 0:
                        break

                    pool_models = [entry for entry in day_schedule if model_to_category[entry['Model']] == category]

                    boost_count, available_slots = distribute_boost_in_pool(model_name, boost_count, pool_models, schedule_dict, available_slots)

                if boost_count < random_entry['Boost Count']:
                    random_entry['Boost Count'] -= 1  # Уменьшаем счетчик бустов для текущей модели

                if random_entry['Boost Count'] == 0:
                    subpool.remove(random_entry)
                    # Обновляем оригинальный boost_pool
                    for entry in boost_pool:
                        if entry['Model'] == random_entry['Model']:
                            entry['Boost Count'] -= 1

        remaining_boosts.extend(
            {
                'Model': entry['Model'],
                'Day': os.path.basename(schedule_file_path).split('_')[0].capitalize(),
                'Remaining Boost Count': entry['Boost Count']
            }
            for entry in boost_pool
        )

        with open(schedule_file_path, 'w') as f:
            json.dump(day_schedule, f, indent=4)

        with open(boost_file_path, 'w') as f:
            json.dump(day_boost, f, indent=4)

        print(f"available_slots - {available_slots}")

        return remaining_boosts

    # Основная обработка по дням недели
    remaining_boosts_all_days = []

    for day in ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']:
        schedule_file = os.path.join(schedule_dir, f'{day}_schedule.json')
        boost_file = os.path.join(boost_dir, f'{day}_boost.json')
        remaining_boosts = distribute_boosts(schedule_file, boost_file)
        remaining_boosts_all_days.extend(remaining_boosts)

    # Открываем или создаем Excel файл
    wb = openpyxl.load_workbook(excel_file_path)
    boost_sheet = wb['Boost']

    # Записываем данные по остаткам бустов обратно в лист Boost
    days_list = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']

    for boost_entry in remaining_boosts_all_days:
        day = boost_entry['Day'].upper()
        day_col = days_list.index(day) + 1  # Находим индекс дня недели
        model_name = boost_entry['Model']
        remaining_boost_count = boost_entry['Remaining Boost Count']

        # Найти строку с нужной моделью
        for row in boost_sheet.iter_rows(min_row=2, max_row=boost_sheet.max_row):
            if row[0].value == model_name:
                # Записать остаток буста в нужную колонку
                row[day_col].value = remaining_boost_count
                break

    # Сохраняем файл Excel
    wb.save(excel_file_path)
    print(f"Данные по остаткам бустов сохранены в файл: {excel_file_path}")

