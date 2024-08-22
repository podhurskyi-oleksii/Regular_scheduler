import json
import random
import os
import openpyxl


def fill_priority_boost():
    # Пути к основным файлам
    models_data_path = 'output_data/models_data.json'
    blacklist_file_path = 'output_data/blacklist_data.json'
    schedule_dir = 'output_data/days_split'
    boost_dir = 'output_data/priority_boost_days_split'
    excel_file_path = 'REGULAR_SCHEDULE.xlsx'

    # Загрузка общих данных
    with open(models_data_path, 'r') as f:
        models_data = json.load(f)

    with open(blacklist_file_path, 'r') as f:
        blacklist_data = json.load(f)

    # Преобразование блокировок для удобства поиска
    blocked_dict = {item['Model']: item['Blocked'] for item in blacklist_data}

    # Создание общего пула моделей
    all_models = [model['Model'] for model in models_data]

    # Функция для проверки блокировок
    def is_blocked(model, target_model):
        blocked_models_for_target = blocked_dict.get(target_model, [])
        return model in blocked_models_for_target

    # Функция для распределения буста по всем моделям
    def distribute_boost_in_pool(model_name, boost_count, tasks_dict):
        attempts = 0
        max_attempts = len(all_models) // 2  # Количество попыток: количество моделей в пуле деленое на 2
        used_models = {model: tasks_dict[model].count(model_name) for model in all_models}

        while boost_count > 0 and attempts < max_attempts:
            available_models = [model for model in all_models if
                                used_models[model] < 2]  # Исключаем модели, где уже 2 буста

            random.shuffle(available_models)
            slots_filled = False

            for target_model in available_models:
                if (
                        target_model != model_name  # Не бустит саму себя
                        and not is_blocked(model_name, target_model)  # Не находится в блоке
                ):
                    # Найти первый свободный слот для буста
                    for i, task in enumerate(tasks_dict[target_model]):
                        if task == 'BOOST':
                            tasks_dict[target_model][i] = f"{model_name}_priority"
                            used_models[target_model] += 1
                            boost_count -= 1
                            slots_filled = True
                            break
                    if boost_count == 0:
                        break

            if not slots_filled:
                attempts += 1
            else:
                attempts = 0  # Сбрасываем счетчик, если хотя бы один слот был заполнен

            if boost_count > 0 and attempts >= max_attempts:
                # Дополнительный проход по оставшимся моделям без рандома
                print(f"Начинается дополнительный проход для модели {model_name}. Оставшийся буст: {boost_count}")
                for target_model in available_models:
                    if (
                            target_model != model_name  # Не бустит саму себя
                            and not is_blocked(model_name, target_model)  # Не находится в блоке
                            and used_models[target_model] < 3  # Не больше 3 раз для одной модели
                    ):
                        # Найти первый свободный слот для буста
                        for i, task in enumerate(tasks_dict[target_model]):
                            if task == 'BOOST':
                                tasks_dict[target_model][i] = f"{model_name}_priority"
                                used_models[target_model] += 1
                                boost_count -= 1
                                if boost_count == 0:
                                    break
                    if boost_count == 0:
                        break

                # Если после дополнительного прохода буст не удалось распределить
                if boost_count > 0:
                    print(f"Не удалось распределить весь буст для модели {model_name}. Оставшийся буст: {boost_count}")
                    break

        return boost_count

    # Основная функция для распределения бустов по дням недели
    def distribute_boosts(schedule_file_path, boost_file_path):
        with open(schedule_file_path, 'r') as f:
            day_schedule = json.load(f)

        with open(boost_file_path, 'r') as f:
            day_boost = json.load(f)

        tasks_dict = {entry['Model']: entry['Tasks'] for entry in day_schedule}

        remaining_boosts = []

        for boost_entry in day_boost:
            model_name = boost_entry['Model']
            boost_count = boost_entry['Boost Count']

            remaining_boost_count = distribute_boost_in_pool(model_name, boost_count, tasks_dict)

            remaining_boosts.append({
                'Model': model_name,
                'Day': os.path.basename(schedule_file_path).split('_')[0].capitalize(),
                'Remaining Boost Count': remaining_boost_count
            })

        # Сохранение обновленного расписания и данных бустов
        with open(schedule_file_path, 'w') as f:
            json.dump(day_schedule, f, indent=4)

        with open(boost_file_path, 'w') as f:
            json.dump(day_boost, f, indent=4)

        return remaining_boosts

    # Открываем или создаем Excel файл
    wb = openpyxl.load_workbook(excel_file_path)
    priority_sheet = wb['PRIORITY']

    # Цикл по дням недели для распределения бустов
    remaining_boosts_all_days = []

    days_of_week = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']

    for day in days_of_week:
        schedule_file = os.path.join(schedule_dir, f'{day}_schedule.json')
        boost_file = os.path.join(boost_dir, f'{day}_boost.json')
        remaining_boosts = distribute_boosts(schedule_file, boost_file)
        remaining_boosts_all_days.extend(remaining_boosts)

    # Записываем данные по остаткам бустов обратно в лист PRIORITY
    days_list = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']

    for boost_entry in remaining_boosts_all_days:
        day = boost_entry['Day'].upper()
        day_col = days_list.index(day) + 1  # Находим индекс дня недели
        model_name = boost_entry['Model']
        remaining_boost_count = boost_entry['Remaining Boost Count']

        # Найти строку с нужной моделью
        for row in priority_sheet.iter_rows(min_row=2, max_row=priority_sheet.max_row):
            if row[0].value == model_name:
                # Записать остаток буста в нужную колонку
                row[day_col].value = remaining_boost_count
                break

    # Сохраняем файл Excel
    wb.save(excel_file_path)
    print(f"Данные по остаткам бустов сохранены в файл: {excel_file_path}")
